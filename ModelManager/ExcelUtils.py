"""
DataDebugger
Class to handle errors and print debug information
"""
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path
from DataManage.Utils.Utils import setup_logger, get_logger, detect_delimiter
import os
import glob
import re
import chardet

"""
Class to work with excel files
Read/Write files

"""


class DataExcel():
    # files names, path, data and status here
    file_names = dict()
    file_data = dict()
    file_errors = dict()
    path_name = Path('../')
    path_data = Path('../data')
    logex = ''
    error_file = False
    error_text = ''

    def __init__(self):
        self.logex = setup_logger('DataExcel', 'data.log')

    # basic fucntions to work with files
    def set_working_path(self, path_name='../'):
        """
        Set working path to work with files

        """
        self.path_name = Path(path_name)
        self.path_data = Path(self.path_data)
        self.logex.info("Set path to: " + str(self.path_name))
        return (self.path_name, self.path_data)

    # load data from Excel
    def load_excel(self, file_name, sheet_name, skip_row=0):
        # path = pth.Path(pathName) # стали использовать try-except вместо  if (path.exists() and path.is_file()):
        try:
            # try to open an existing workbook, mode='a' - append
            data_var = pd.read_excel(file_name, sheet_name, skiprows=skip_row)  # utf-8 - cp-1251
            self.error_file = False
            self.error_text = None
            self.logex.info("Read excel file: " + str(file_name) + ' sheet: ' + str(sheet_name))
        except FileNotFoundError:
            data_var = None
            self.error_file = True
            self.error_text = "Error: file not found: " + str(file_name)
            self.logex.error(self.error_text)
        except PermissionError:
            data_var = None
            self.error_file = True
            self.error_text = "Error: could not open file, permission denied: " + str(file_name)
            self.logex.error(self.error_text)
        except Exception as e:
            data_var = None
            self.error_file = True
            self.error_text = "Error: something happened: " + str(e)
            self.logex.error(self.error_text)
        return data_var

        # convert dates, add to dictiony

    def load_data(self, file_name, file_sheet, date_form="%Y-%m-%d"):
        file_data = self.load_excel(file_name, file_sheet)
        if ~self.error_file:
            file_data["date"] = pd.to_datetime(file_data.iloc[:, 0], format=date_form)
            file_data = file_data.set_index("date")
        else:
            file_data = None
        self.file_data[file_name] = file_data
        self.file_errors[file_name] = self.error_text
        return file_data

        # save data to excel

    def load_file_type(self, file_path, params):
        """Fast load files with different types
        :param params: 'excel_sheet' - name of excel sheet;  'csv_head' - header in csv;
        'csv_sep' - separator for csv;
        :param file_path: pass to file
        :param file_type: type of the file,  excel, csv or feather
        """

        file_extension = os.path.splitext(file_path)[-1].lower()
        # Read the file based on its extension
        try:
            if file_extension == '.csv':
                # Detect the file encoding
                encode, delime = detect_delimiter(file_path)
                # Read the CSV file with the detected encoding
                df = pd.read_csv(file_path, encoding=encode, sep=delime,
                                 header=params['csv_head'])
                self.logex.info("Read csv file: " + str(file_path))
            elif file_extension in ('.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.ods', '.odt'):
                df = pd.read_excel(file_path, params['excel_sheet'], engine='openpyxl')
                self.logex.info("Read excel file: " + str(file_path) + ' sheet: ' + str(params['excel_sheet']))
            elif file_extension in ('.ftr', '.f'):
                df = pd.read_feather(file_path)
                self.logex.info("Read csv file: " + str(file_path))
            elif file_extension == '.json':
                df = pd.read_json(file_path, lines= params['json_lines'])
                self.logex.info("Read json file: " + str(file_path))
            else:
                self.error_file = True
                self.error_text = "Error: unexpected file format: " + str(file_path)
                self.logex.error(self.error_text)
                df = None  # Skip the file if it has an unrecognized extension
        except Exception as e:
            df = None
            self.error_file = True
            self.error_text = "Error: " + str(e)
            self.logex.error(self.error_text)
        return df

    def save_excel(self, file_name, data_set, file_sheet='Sheet1', st_row=0):
        try:
            # try to open an existing workbook, mode='a' - append
            writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a')
            writer.book = load_workbook(file_name)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            # search for work_sheet
            if file_sheet in writer.book.sheetnames:
                data_set.to_excel(writer, sheet_name=file_sheet, startrow=st_row, index=True, if_sheet_exists='replace')
                # replace data in workseet
                # data_set.to_excel(writer, sheet_name = 'data', index=True)  #original cpi mom
                # if not create new sheet
            else:
                data_set.to_excel(writer, sheet_name=file_sheet, startrow=st_row, index=True)
            writer.save()
            writer.close()
        # if could not open file, then create new one
        except FileNotFoundError:
            self.error_file = True
            self.error_text = "Error: file not found: " + str(file_name)
            self.logex.error(self.error_text)
            data_set.to_excel(file_name, sheet_name=file_sheet, startrow=st_row)
        except PermissionError:
            self.error_file = True
            self.error_text = "Error: could not open file, permission denied: " + str(file_name)
            self.logex.error(self.error_text)

    def excel_saver(self, filePath, data_df, sheet_name):
        # read the existing sheets so that openpyxl won't create a new one later
        try:
            book = load_workbook(filePath)
            writer = pd.ExcelWriter(filePath, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            data_df.to_excel(writer, sheet_name, startrow=0, startcol=0)
            writer.save()
            self.logex.info("Save excel file: " + str(filePath) + ' sheet: ' + str(sheet_name))
        except:
            self.logex.error("Save excel file Error: " + str(filePath) + ' sheet: ' + str(sheet_name))

